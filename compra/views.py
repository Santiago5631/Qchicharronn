# compra/views.py
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.shortcuts import render, get_object_or_404
from .models import *
from .forms import CompraForm
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import io

from usuario.permisos import RolRequeridoMixin, rol_requerido, SOLO_ADMIN


class CompraListView(RolRequeridoMixin, ListView):
    """Solo administradores pueden ver compras."""
    roles_permitidos = SOLO_ADMIN
    model = Compra
    template_name = 'modulos/compra.html'
    context_object_name = 'compras'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Listado de Compras'
        return context


class CompraCreateView(RolRequeridoMixin, CreateView):
    """Solo administradores pueden crear compras."""
    roles_permitidos = SOLO_ADMIN
    model = Compra
    template_name = 'forms/formulario_crear.html'
    form_class = CompraForm
    success_url = '/apps/compras/listar/'

    def form_valid(self, form):
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Compra'
        context['modulo'] = "compra"
        return context


class CompraUpdateView(RolRequeridoMixin, UpdateView):
    """Solo administradores pueden editar compras."""
    roles_permitidos = SOLO_ADMIN
    model = Compra
    template_name = 'forms/formulario_actualizacion.html'
    fields = ['producto', 'fecha', 'precio', 'proveedor', 'cantidad', 'unidad']
    success_url = '/apps/compras/listar/'

    def form_valid(self, form):
        return super().form_valid(form)


class CompraDeleteView(RolRequeridoMixin, DeleteView):
    """Solo administradores pueden eliminar compras."""
    roles_permitidos = SOLO_ADMIN
    model = Compra
    template_name = 'forms/confirmar_eliminacion.html'
    success_url = '/apps/compras/listar/'

    def post(self, request, *args, **kwargs):
        from django.http import JsonResponse
        try:
            self.get_object().delete()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ──────────────────────────────────────────────
# EXPORTAR A EXCEL — Solo administradores
# ──────────────────────────────────────────────
@rol_requerido(*SOLO_ADMIN)
def exportar_compras_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    encabezados = ['ID', 'Producto', 'Fecha', 'Precio', 'Proveedor', 'Cantidad', 'Unidad']
    ws.append(encabezados)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(encabezados) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    compras = Compra.objects.select_related('producto', 'proveedor').all().order_by('-fecha')
    total_precio = 0

    for c in compras:
        precio = c.precio or 0
        total_precio += precio
        ws.append([
            c.id_factura or '',
            c.producto.nombre if c.producto else 'Sin producto',
            c.fecha.strftime('%d/%m/%Y') if c.fecha else '',
            f"${precio:,.2f}",
            str(c.proveedor) if c.proveedor else 'Sin proveedor',
            c.cantidad or 0,
            str(c.unidad) if c.unidad else '',
        ])

    ws.append([])
    ws.append(['', '', '', f'Total General:', f"${total_precio:,.2f}", '', ''])

    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions[get_column_letter(2)].width = 30

    filename = f"compras_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
# EXPORTAR A PDF — Solo administradores
# ──────────────────────────────────────────────
@rol_requerido(*SOLO_ADMIN)
def exportar_compras_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph(
        f"Listado de Compras<br/><small>Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}</small>",
        styles['Title']
    )
    elements.append(titulo)
    elements.append(Spacer(1, 20))

    data = [['ID', 'Producto', 'Fecha', 'Precio', 'Proveedor', 'Cantidad', 'Unidad']]

    compras = Compra.objects.select_related('producto', 'proveedor').all().order_by('-fecha')
    total_precio = 0

    for c in compras:
        precio = c.precio or 0
        total_precio += precio
        data.append([
            c.id_factura or '-',
            c.producto.nombre if c.producto else 'Sin producto',
            c.fecha.strftime('%d/%m/%Y') if c.fecha else '-',
            f"${precio:,.2f}",
            str(c.proveedor) if c.proveedor else 'Sin proveedor',
            str(c.cantidad or ''),
            str(c.unidad or ''),
        ])

    data.append(['', '', 'TOTAL:', f"${total_precio:,.2f}", '', '', ''])

    table = Table(data, colWidths=[50, 120, 70, 70, 100, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"compras_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@rol_requerido(*SOLO_ADMIN)
def exportar_compras_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
    )
    from reportlab.lib.units import mm
    import io
    from django.utils import timezone

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=22*mm,
        leftMargin=22*mm,
        topMargin=32*mm,
        bottomMargin=24*mm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Estilos personalizados
    title_style = ParagraphStyle(
        name='ElegantTitle',
        fontSize=20,
        textColor=colors.HexColor('#1a3c6d'),
        alignment=1,  # center
        spaceAfter=8,
        leading=24,
    )

    subtitle_style = ParagraphStyle(
        name='Subtitle',
        fontSize=10.5,
        textColor=colors.HexColor('#5a7184'),
        alignment=1,
        spaceAfter=28,
    )

    header_cell_style = ParagraphStyle(
        name='HeaderCell',
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.white,
        alignment=1,  # center
        leading=12,
    )

    normal_cell_style = ParagraphStyle(
        name='NormalCell',
        fontSize=9.5,
        leading=11.5,
        alignment=0,  # left por defecto
    )

    price_cell_style = ParagraphStyle(
        name='PriceCell',
        parent=normal_cell_style,
        alignment=2,  # right
    )

    total_style = ParagraphStyle(
        name='TotalStyle',
        fontSize=11,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1a3c6d'),
        alignment=2,  # right
    )

    # Título principal
    elements.append(Paragraph("LISTADO DE COMPRAS", title_style))

    # Línea decorativa dorada debajo del título
    elements.append(Spacer(1, 6))
    elements.append(HRFlowable(
        width="60%",
        thickness=1.2,
        lineCap='round',
        color=colors.HexColor('#d4af37'),  # dorado elegante
        spaceBefore=4,
        spaceAfter=18,
        hAlign='CENTER'
    ))

    # Fecha de generación
    fecha = timezone.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Generado el {fecha}", subtitle_style))

    elements.append(Spacer(1, 18))

    # ── Tabla ────────────────────────────────────────────────────────────────
    data = [[
        Paragraph("ID", header_cell_style),
        Paragraph("Producto", header_cell_style),
        Paragraph("Fecha", header_cell_style),
        Paragraph("Precio", header_cell_style),
        Paragraph("Proveedor", header_cell_style),
        Paragraph("Cantidad", header_cell_style),
        Paragraph("Unidad", header_cell_style),
    ]]

    compras = Compra.objects.select_related('producto').order_by('-fecha')
    total_precio = 0.0

    for compra in compras:
        precio = float(compra.precio or 0)
        total_precio += precio

        unidad_txt = str(compra.unidad) if compra.unidad else ""
        if len(unidad_txt) > 55:
            unidad_txt = unidad_txt[:52] + "..."

        data.append([
            Paragraph(str(compra.id_factura or "—"), normal_cell_style),
            Paragraph(str(compra.producto.nombre if compra.producto else "—"), normal_cell_style),
            Paragraph(compra.fecha.strftime("%d/%m/%Y") if compra.fecha else "—", normal_cell_style),
            Paragraph(f"S/ {precio:,.2f}", price_cell_style),
            Paragraph(str(compra.proveedor) if compra.proveedor else "—", normal_cell_style),
            Paragraph(f"{compra.cantidad or 0:g}", normal_cell_style),
            Paragraph(unidad_txt, normal_cell_style),
        ])

    # Fila total
    data.append([
        Paragraph("TOTAL GENERAL", total_style),
        Paragraph("", normal_cell_style),
        Paragraph("", normal_cell_style),
        Paragraph("", normal_cell_style),
        Paragraph("", normal_cell_style),
        Paragraph("", normal_cell_style),
        Paragraph(f"S/ {total_precio:,.2f}", total_style),
    ])

    # Anchos de columnas (ajustados para que quepa bien en A4)
    col_widths = [48, 138, 68, 78, 105, 58, 98]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING',    (0, 0), (-1, 0), 10),

        # Bordes redondeados simulados (solo visual)
        ('BOX',        (0, 0), (-1, -1), 0.8, colors.HexColor('#0d6efd')),
        ('GRID',       (0, 1), (-1, -1), 0.4, colors.HexColor('#d0e0ff')),

        # Alineaciones por columna
        ('ALIGN',      (0, 1), (0, -1), 'CENTER'),   # ID
        ('ALIGN',      (2, 1), (2, -1), 'CENTER'),   # Fecha
        ('ALIGN',      (3, 1), (3, -1), 'RIGHT'),    # Precio
        ('ALIGN',      (5, 1), (5, -1), 'CENTER'),   # Cantidad
        ('ALIGN',      (6, 1), (6, -2), 'LEFT'),     # Unidad

        # Fondo filas alternas muy sutil
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fbff')]),

        # Total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e7f1ff')),
        ('SPAN',       (0, -1), (5, -1)),  # TOTAL GENERAL ocupa hasta Cantidad
        ('ALIGN',      (6, -1), (6, -1), 'RIGHT'),
        ('FONTSIZE',   (0, -1), (-1, -1), 10.8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
        ('TOPPADDING',    (0, -1), (-1, -1), 12),
    ]))

    elements.append(table)

    # Pie de página
    elements.append(Spacer(1, 36))
    elements.append(Paragraph(
        "Reporte generado por el sistema de compras",
        ParagraphStyle(
            name='Footer',
            fontSize=8.5,
            textColor=colors.HexColor('#6c757d'),
            alignment=1,
        )
    ))

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"listado_compras_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response