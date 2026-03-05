# informe/views.py
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Sum
from django.utils import timezone
import datetime
import io

from venta.models import Venta
from compra.models import Compra
from inventario.models import InventarioDiario
from menu.models import Menu
from .models import Informe

from usuario.permisos import RolRequeridoMixin, rol_requerido, SOLO_ADMIN


# ──────────────────────────────────────────────
# HELPER: construir queryset según tipo y fechas
# ──────────────────────────────────────────────
def _obtener_resultados(tipo, fecha_inicio, fecha_fin, contexto):
    """
    Rellena `contexto` con resultados, cantidad, total y mensaje.
    Devuelve True si hubo resultados, False si no.
    """
    try:
        if tipo == 'ventas':
            qs = Venta.objects.filter(
                fecha_venta__date__gte=fecha_inicio,
                fecha_venta__date__lte=fecha_fin,
            ).select_related('pedido').order_by('-fecha_venta')
            contexto['resultados'] = qs
            contexto['cantidad']   = qs.count()
            total_sum = qs.aggregate(s=Sum('total'))['s']
            contexto['total']  = total_sum or 0
            if contexto['cantidad']:
                contexto['mensaje'] = (
                    f"Se encontraron {contexto['cantidad']} ventas — "
                    f"Total: ${contexto['total']:,.2f}"
                )
            else:
                contexto['sin_resultados'] = True
                contexto['mensaje'] = (
                    "No se encontraron ventas en el rango de fechas seleccionado."
                )

        elif tipo == 'compras':
            qs = Compra.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
            ).select_related('producto', 'proveedor').order_by('-fecha')
            # calcular subtotal en memoria (campo dinámico)
            filas = list(qs)
            for c in filas:
                c.subtotal = c.precio * c.cantidad
            contexto['resultados'] = filas
            contexto['cantidad']   = len(filas)
            contexto['total']      = sum(c.subtotal for c in filas)
            if filas:
                contexto['mensaje'] = (
                    f"Se encontraron {contexto['cantidad']} compras — "
                    f"Total: ${contexto['total']:,.2f}"
                )
            else:
                contexto['sin_resultados'] = True
                contexto['mensaje'] = (
                    "No se encontraron compras en el rango de fechas seleccionado."
                )

        elif tipo == 'pedidos':
            from pedido.models import Pedido
            qs = Pedido.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
            ).prefetch_related('detalles__menu').order_by('-fecha')
            filas = list(qs)
            for pedido in filas:
                total = 0
                detalles_list = []
                for detalle in pedido.detalles.all():
                    if hasattr(detalle.menu, 'get_precio_final'):
                        precio = detalle.menu.get_precio_final()
                    elif hasattr(detalle.menu, 'precio_base'):
                        precio = detalle.menu.precio_base
                    else:
                        precio = 0
                    subtotal = precio * detalle.cantidad
                    total += subtotal
                    detalles_list.append({
                        'menu': detalle.menu.nombre,
                        'cantidad': detalle.cantidad,
                        'precio': precio,
                        'subtotal': subtotal,
                    })
                pedido.total_calculado = total
                pedido.detalles_list   = detalles_list
            contexto['resultados'] = filas
            contexto['cantidad']   = len(filas)
            contexto['total']      = sum(p.total_calculado for p in filas)
            if filas:
                contexto['mensaje'] = (
                    f"Se encontraron {contexto['cantidad']} pedidos — "
                    f"Total: ${contexto['total']:,.2f}"
                )
            else:
                contexto['sin_resultados'] = True
                contexto['mensaje'] = (
                    "No se encontraron pedidos en el rango de fechas seleccionado."
                )

        elif tipo == 'inventario':
            qs = InventarioDiario.objects.filter(
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
            ).order_by('-fecha')
            contexto['resultados'] = qs
            contexto['cantidad']   = qs.count()
            contexto['abiertos']   = qs.filter(estado='abierto').count()
            contexto['cerrados']   = qs.filter(estado='cerrado').count()
            if contexto['cantidad']:
                contexto['mensaje'] = (
                    f"{contexto['cantidad']} registros de inventario encontrados"
                )
            else:
                contexto['sin_resultados'] = True
                contexto['mensaje'] = (
                    "No hay registros de inventario en el rango de fechas seleccionado."
                )
        else:
            contexto['mensaje'] = "Tipo de reporte no reconocido."

    except Exception as e:
        contexto['mensaje'] = f"Error al consultar: {str(e)}"


# ──────────────────────────────────────────────
# VISTA PRINCIPAL DE REPORTES — Solo administradores
# ──────────────────────────────────────────────
@rol_requerido(*SOLO_ADMIN)
def informe_list(request):
    tipo             = request.GET.get('tipo', 'ventas')
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')
    consultado       = bool(fecha_inicio_str and fecha_fin_str)   # ¿vino del form?

    contexto = {
        'titulo':         'Consulta de Reportes',
        'tipo_actual':    tipo,
        'fecha_inicio':   fecha_inicio_str,
        'fecha_fin':      fecha_fin_str,
        'mensaje':        "Selecciona un tipo y un rango de fechas para consultar.",
        'resultados':     None,
        'total':          0,
        'cantidad':       0,
        'sin_resultados': False,
        'consultado':     consultado,
        'informes':       Informe.objects.all().order_by('-fecha_creacion'),
    }

    if not consultado:
        return render(request, 'modulos/informe.html', contexto)

    # Parsear fechas
    fecha_inicio = fecha_fin = None
    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    except Exception:
        contexto['mensaje'] = "Formato de fecha de inicio inválido."
        return render(request, 'modulos/informe.html', contexto)

    try:
        fecha_fin = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except Exception:
        contexto['mensaje'] = "Formato de fecha de fin inválido."
        return render(request, 'modulos/informe.html', contexto)

    if fecha_inicio > fecha_fin:
        contexto['mensaje'] = "La fecha de inicio no puede ser mayor que la fecha final."
        return render(request, 'modulos/informe.html', contexto)

    _obtener_resultados(tipo, fecha_inicio, fecha_fin, contexto)
    return render(request, 'modulos/informe.html', contexto)


# ──────────────────────────────────────────────
# CRUD de Informes — Solo administradores
# ──────────────────────────────────────────────
class InformeListView(RolRequeridoMixin, ListView):
    roles_permitidos = SOLO_ADMIN
    model            = Informe
    template_name    = 'modulos/informe.html'
    context_object_name = 'informes'


class InformeCreateView(RolRequeridoMixin, CreateView):
    roles_permitidos = SOLO_ADMIN
    model   = Informe
    fields  = ['titulo', 'descripcion', 'tipo', 'fecha_inicio', 'fecha_fin']
    template_name = 'forms/formulario_crear.html'
    success_url   = reverse_lazy('apl:informe:informe_list')

    def form_valid(self, form):
        messages.success(self.request, "Informe creado correctamente")
        return super().form_valid(form)


class InformeUpdateView(RolRequeridoMixin, UpdateView):
    roles_permitidos = SOLO_ADMIN
    model   = Informe
    fields  = ['titulo', 'descripcion', 'tipo', 'fecha_inicio', 'fecha_fin']
    template_name = 'forms/formulario_actualizacion.html'
    success_url   = reverse_lazy('apl:informe:informe_list')

    def form_valid(self, form):
        messages.success(self.request, "Informe actualizado correctamente")
        return super().form_valid(form)


class InformeDeleteView(RolRequeridoMixin, DeleteView):
    roles_permitidos = SOLO_ADMIN
    model       = Informe
    success_url = reverse_lazy('apl:informe:informe_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return JsonResponse({"status": "ok"})


# ──────────────────────────────────────────────
# EXPORTAR EXCEL
# ──────────────────────────────────────────────
@rol_requerido(*SOLO_ADMIN)
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tipo             = request.GET.get('tipo', 'ventas')
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')

    if not (fecha_inicio_str and fecha_fin_str):
        return HttpResponse("Selecciona un rango de fechas antes de exportar.", status=400)

    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin    = datetime.datetime.strptime(fecha_fin_str,    '%Y-%m-%d').date()
    except Exception:
        return HttpResponse("Fechas inválidas.", status=400)

    ctx = {'resultados': None, 'cantidad': 0, 'total': 0,
           'sin_resultados': False, 'mensaje': ''}
    _obtener_resultados(tipo, fecha_inicio, fecha_fin, ctx)

    wb = openpyxl.Workbook()
    ws = wb.active

    # ── Estilos ──
    naranja    = PatternFill("solid", fgColor="E85A1E")
    gris_head  = PatternFill("solid", fgColor="F2F2F2")
    fila_par   = PatternFill("solid", fgColor="FFF8F5")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    bold_dark  = Font(bold=True, size=10)
    normal     = Font(size=10)
    center     = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # ── Título del reporte ──
    ws.merge_cells('A1:F1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"Q'CHICHARRÓN — Reporte de {tipo.capitalize()}"
    titulo_cell.font      = Font(bold=True, color="FFFFFF", size=14)
    titulo_cell.fill      = naranja
    titulo_cell.alignment = center

    ws.merge_cells('A2:F2')
    sub = ws['A2']
    sub.value     = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}    |    Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font      = Font(size=9, italic=True, color="777777")
    sub.alignment = center
    sub.fill      = PatternFill("solid", fgColor="FFF3EE")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    # ── Cabeceras y datos según tipo ──
    HEADERS = {
        'ventas':     ['Fecha', 'N° Pedido', 'Total ($)', 'Método de Pago', 'Estado', 'Administrador'],
        'compras':    ['Fecha', 'Producto', 'Cantidad', 'Precio ($)', 'Subtotal ($)', 'Proveedor'],
        'pedidos':    ['ID', 'Fecha', 'Mesa', 'Ítems', 'Total ($)', 'Estado'],
        'inventario': ['Fecha', 'Estado', 'Abiertos', 'Cerrados', '', ''],
    }
    headers = HEADERS.get(tipo, [])

    # fila 3 = cabeceras
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = bold_white
        cell.fill      = PatternFill("solid", fgColor="C0392B")
        cell.alignment = center
        cell.border    = thin_border
    ws.row_dimensions[3].height = 20

    # ── Filas de datos ──
    fila = 4
    resultados = ctx.get('resultados') or []

    if tipo == 'ventas':
        for i, v in enumerate(resultados):
            fill = fila_par if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            datos = [
                v.fecha_venta.strftime('%d/%m/%Y') if hasattr(v.fecha_venta, 'strftime') else str(v.fecha_venta),
                str(v.pedido.id) if v.pedido else '—',
                float(v.total),
                v.get_metodo_pago_display() if hasattr(v, 'get_metodo_pago_display') else '—',
                v.get_estado_display() if hasattr(v, 'get_estado_display') else '—',
                str(v.admin) if hasattr(v, 'admin') and v.admin else '—',
            ]
            for col_idx, val in enumerate(datos, 1):
                cell = ws.cell(row=fila, column=col_idx, value=val)
                cell.font = normal; cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(horizontal='right' if col_idx == 3 else 'left', vertical='center')
            fila += 1

    elif tipo == 'compras':
        for i, c in enumerate(resultados):
            fill = fila_par if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            datos = [
                c.fecha.strftime('%d/%m/%Y') if hasattr(c.fecha, 'strftime') else str(c.fecha),
                str(c.producto.nombre) if hasattr(c.producto, 'nombre') else str(c.producto),
                float(c.cantidad),
                float(c.precio),
                float(c.subtotal),
                str(c.proveedor.nombre) if hasattr(c.proveedor, 'nombre') else str(c.proveedor),
            ]
            for col_idx, val in enumerate(datos, 1):
                cell = ws.cell(row=fila, column=col_idx, value=val)
                cell.font = normal; cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='right' if col_idx in (3, 4, 5) else 'left',
                    vertical='center'
                )
            fila += 1

    elif tipo == 'pedidos':
        for i, p in enumerate(resultados):
            fill = fila_par if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            items_str = ', '.join(
                f"{d['menu']} x{d['cantidad']}" for d in getattr(p, 'detalles_list', [])
            )
            datos = [
                f"#{p.id}",
                p.fecha.strftime('%d/%m/%Y %H:%M') if hasattr(p.fecha, 'strftime') else str(p.fecha),
                str(p.mesa) if p.mesa else 'Para llevar',
                items_str,
                float(getattr(p, 'total_calculado', 0)),
                p.get_estado_display() if hasattr(p, 'get_estado_display') else '—',
            ]
            for col_idx, val in enumerate(datos, 1):
                cell = ws.cell(row=fila, column=col_idx, value=val)
                cell.font = normal; cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='right' if col_idx == 5 else 'left',
                    vertical='center', wrap_text=(col_idx == 4)
                )
            fila += 1

    elif tipo == 'inventario':
        for i, inv in enumerate(resultados):
            fill = fila_par if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            datos = [
                inv.fecha.strftime('%d/%m/%Y') if hasattr(inv.fecha, 'strftime') else str(inv.fecha),
                inv.estado.capitalize() if hasattr(inv, 'estado') else '—',
                '', '', '', '',
            ]
            for col_idx, val in enumerate(datos, 1):
                cell = ws.cell(row=fila, column=col_idx, value=val)
                cell.font = normal; cell.fill = fill; cell.border = thin_border
            fila += 1

    # ── Fila de totales ──
    if ctx.get('total') and fila > 4:
        ws.cell(row=fila, column=1, value='TOTAL').font = bold_dark
        total_col = {'ventas': 3, 'compras': 5, 'pedidos': 5}.get(tipo)
        if total_col:
            tc = ws.cell(row=fila, column=total_col, value=float(ctx['total']))
            tc.font = Font(bold=True, size=10, color="C0392B")
            tc.alignment = Alignment(horizontal='right')
            tc.fill = PatternFill("solid", fgColor="FFF3EE")

    # ── Ancho de columnas ──
    anchos = {
        'ventas':     [14, 12, 14, 18, 12, 18],
        'compras':    [14, 22, 12, 14, 14, 20],
        'pedidos':    [10, 18, 14, 35, 14, 14],
        'inventario': [14, 14, 12, 12, 10, 10],
    }
    for idx, ancho in enumerate(anchos.get(tipo, []), 1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    # ── Sin datos ──
    if not resultados:
        ws.merge_cells(f'A4:F4')
        cell_nd = ws['A4']
        cell_nd.value     = f"No se encontraron registros de {tipo} en el período seleccionado."
        cell_nd.font      = Font(italic=True, color="999999")
        cell_nd.alignment = center

    # ── Respuesta HTTP ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"reporte_{tipo}_{fecha_inicio_str}_al_{fecha_fin_str}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ──────────────────────────────────────────────
# EXPORTAR PDF
# ──────────────────────────────────────────────
@rol_requerido(*SOLO_ADMIN)
def exportar_pdf(request):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    tipo             = request.GET.get('tipo', 'ventas')
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')

    if not (fecha_inicio_str and fecha_fin_str):
        return HttpResponse("Selecciona un rango de fechas antes de exportar.", status=400)

    try:
        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin    = datetime.datetime.strptime(fecha_fin_str,    '%Y-%m-%d').date()
    except Exception:
        return HttpResponse("Fechas inválidas.", status=400)

    ctx = {'resultados': None, 'cantidad': 0, 'total': 0,
           'sin_resultados': False, 'mensaje': ''}
    _obtener_resultados(tipo, fecha_inicio, fecha_fin, ctx)

    # ── Estilos ──
    buffer   = io.BytesIO()
    doc      = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles   = getSampleStyleSheet()
    NARANJA  = colors.HexColor('#E85A1E')
    ROJO     = colors.HexColor('#C0392B')
    GRIS_CLR = colors.HexColor('#F2F2F2')
    FILA_PAR = colors.HexColor('#FFF8F5')

    estilo_titulo = ParagraphStyle(
        'titulo', parent=styles['Title'],
        fontSize=16, textColor=colors.white,
        alignment=TA_CENTER, spaceAfter=4,
    )
    estilo_sub = ParagraphStyle(
        'sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#555555'),
        alignment=TA_CENTER, spaceAfter=6,
    )
    estilo_total = ParagraphStyle(
        'total', parent=styles['Normal'],
        fontSize=11, textColor=ROJO,
        alignment=TA_RIGHT,
    )
    estilo_vacio = ParagraphStyle(
        'vacio', parent=styles['Normal'],
        fontSize=11, textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER,
    )

    elementos = []

    # ── Encabezado ──
    from reportlab.platypus import Table as RLTable
    header_data = [[
        Paragraph(f"Q'CHICHARRÓN — Reporte de {tipo.capitalize()}", estilo_titulo),
    ]]
    header_table = RLTable(header_data, colWidths=['100%'])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), NARANJA),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 10),
    ]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 0.15*cm))
    elementos.append(Paragraph(
        f"Período: {fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}   |   "
        f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
        estilo_sub
    ))
    elementos.append(Spacer(1, 0.3*cm))

    resultados = ctx.get('resultados') or []

    if not resultados:
        elementos.append(Spacer(1, 1*cm))
        elementos.append(Paragraph(
            f"No se encontraron registros de {tipo} en el período seleccionado.",
            estilo_vacio
        ))
    else:
        # ── Tabla de datos ──
        HEADERS_PDF = {
            'ventas':     ['Fecha', 'N° Pedido', 'Total ($)', 'Método', 'Estado', 'Administrador'],
            'compras':    ['Fecha', 'Producto', 'Cantidad', 'Precio ($)', 'Subtotal ($)', 'Proveedor'],
            'pedidos':    ['ID', 'Fecha', 'Mesa', 'Ítems', 'Total ($)', 'Estado'],
            'inventario': ['Fecha', 'Estado', '', '', '', ''],
        }
        COL_WIDTHS = {
            'ventas':     [3.5*cm, 3*cm, 3.5*cm, 4*cm, 3*cm, 4*cm],
            'compras':    [3.5*cm, 5*cm, 3*cm, 3.5*cm, 3.5*cm, 4.5*cm],
            'pedidos':    [2*cm, 4*cm, 3.5*cm, 8*cm, 3.5*cm, 3*cm],
            'inventario': [4*cm, 4*cm, 3*cm, 3*cm, 3*cm, 3*cm],
        }

        cell_style = ParagraphStyle('cell', fontSize=9, leading=11)
        cell_r     = ParagraphStyle('cellr', fontSize=9, leading=11, alignment=TA_RIGHT)

        filas_tabla = [HEADERS_PDF.get(tipo, [])]

        if tipo == 'ventas':
            for v in resultados:
                filas_tabla.append([
                    v.fecha_venta.strftime('%d/%m/%Y') if hasattr(v.fecha_venta, 'strftime') else str(v.fecha_venta),
                    str(v.pedido.id) if v.pedido else '—',
                    Paragraph(f"${float(v.total):,.2f}", cell_r),
                    v.get_metodo_pago_display() if hasattr(v, 'get_metodo_pago_display') else '—',
                    v.get_estado_display() if hasattr(v, 'get_estado_display') else '—',
                    str(v.admin) if hasattr(v, 'admin') and v.admin else '—',
                ])
        elif tipo == 'compras':
            for c in resultados:
                filas_tabla.append([
                    c.fecha.strftime('%d/%m/%Y') if hasattr(c.fecha, 'strftime') else str(c.fecha),
                    str(c.producto.nombre) if hasattr(c.producto, 'nombre') else str(c.producto),
                    Paragraph(str(float(c.cantidad)), cell_r),
                    Paragraph(f"${float(c.precio):,.2f}", cell_r),
                    Paragraph(f"${float(c.subtotal):,.2f}", cell_r),
                    str(c.proveedor.nombre) if hasattr(c.proveedor, 'nombre') else str(c.proveedor),
                ])
        elif tipo == 'pedidos':
            for p in resultados:
                items_str = '\n'.join(
                    f"• {d['menu']} x{d['cantidad']}"
                    for d in getattr(p, 'detalles_list', [])
                )
                filas_tabla.append([
                    f"#{p.id}",
                    p.fecha.strftime('%d/%m/%Y %H:%M') if hasattr(p.fecha, 'strftime') else str(p.fecha),
                    str(p.mesa) if p.mesa else 'Para llevar',
                    Paragraph(items_str.replace('\n', '<br/>'), cell_style),
                    Paragraph(f"${float(getattr(p, 'total_calculado', 0)):,.2f}", cell_r),
                    p.get_estado_display() if hasattr(p, 'get_estado_display') else '—',
                ])
        elif tipo == 'inventario':
            for inv in resultados:
                filas_tabla.append([
                    inv.fecha.strftime('%d/%m/%Y') if hasattr(inv.fecha, 'strftime') else str(inv.fecha),
                    inv.estado.capitalize() if hasattr(inv, 'estado') else '—',
                    '', '', '', '',
                ])

        # ── Fila de total ──
        if ctx.get('total') and tipo in ('ventas', 'compras', 'pedidos'):
            fila_total = ['', '', '', '', '', '']
            fila_total[0] = 'TOTAL'
            col_total = {'ventas': 2, 'compras': 4, 'pedidos': 4}.get(tipo, 2)
            fila_total[col_total] = Paragraph(
                f"<b>${float(ctx['total']):,.2f}</b>", cell_r
            )
            filas_tabla.append(fila_total)

        tabla = Table(filas_tabla, colWidths=COL_WIDTHS.get(tipo, None), repeatRows=1)

        n_filas = len(filas_tabla)
        estilos_tabla = [
            # Encabezado
            ('BACKGROUND',   (0, 0), (-1, 0), ROJO),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 9),
            ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING',   (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING',(0, 0), (-1, 0), 6),
            # Datos
            ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',     (0, 1), (-1, -1), 8),
            ('TOPPADDING',   (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING',(0, 1), (-1, -1), 4),
            ('LEFTPADDING',  (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            # Bordes
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
            ('LINEBELOW',    (0, 0), (-1, 0), 1.2, ROJO),
            # Filas alternadas
            *[
                ('BACKGROUND', (0, i), (-1, i), FILA_PAR if i % 2 == 0 else colors.white)
                for i in range(1, n_filas - (1 if ctx.get('total') else 0))
            ],
            # Fila total
            *([('BACKGROUND', (0, n_filas-1), (-1, n_filas-1), colors.HexColor('#FFF3EE')),
               ('FONTNAME',   (0, n_filas-1), (-1, n_filas-1), 'Helvetica-Bold'),
               ('LINEABOVE',  (0, n_filas-1), (-1, n_filas-1), 1, ROJO)]
              if ctx.get('total') and tipo in ('ventas', 'compras', 'pedidos') else []),
        ]
        tabla.setStyle(TableStyle(estilos_tabla))
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.4*cm))

        # ── Resumen ──
        elementos.append(Paragraph(
            f"Total de registros: <b>{ctx['cantidad']}</b>"
            + (f"   |   Total general: <b>${float(ctx['total']):,.2f}</b>"
               if ctx.get('total') else ''),
            estilo_total
        ))

    doc.build(elementos)
    buffer.seek(0)

    nombre_archivo = f"reporte_{tipo}_{fecha_inicio_str}_al_{fecha_fin_str}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response